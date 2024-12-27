import os
import random
import time
import aiofiles
from aiohttp.client_exceptions import ConnectionTimeoutError
from openpyxl import Workbook, load_workbook
import asyncio
from aiomultiprocess import Pool
import aiohttp
from bs4 import BeautifulSoup, ResultSet
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


async def get_main_page(url):
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            read = await response.read()
            read = BeautifulSoup(read, "html.parser")
            # print(read)
            pretty = read.prettify()

    async with aiofiles.open('index.html', "w", encoding="utf-8") as f:
        await f.write(pretty)

async def href_from_table():
    try:    
        async with aiofiles.open("index.html", "r", encoding="utf-8") as f:
            html = await f.read()
        soup = BeautifulSoup(html, 'html.parser')
        top = soup.find(id="curr_top")
        ass1 = top.find_all("a")
        top_hrefs = [href.get('href') for href in ass1]
        table = soup.find(id="curr_tab_c")
        body = table.find("tbody")
        ass2 = body.find_all("a")
        hrefs = [href.get("href") for href in ass2]
        # print(len(hrefs))
        # with open("_a_tag.html", "w", encoding='utf-8') as f:
        #     f.write('\n'.join([str(tag) for tag in ass]))
        # for raw in ass:
        #     print(raw)
        async with aiofiles.open("hrefs2.txt", "w", encoding='utf-8') as f:
            await f.write('\n'.join(top_hrefs))
            await f.write('\n'.join(hrefs))
        
        return top_hrefs, hrefs
    except Exception as e:
        print('Error handled: ', e)
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")

        # Automatically resolve the ChromeDriver using Selenium Manager
        driver = webdriver.Chrome(options=chrome_options)

        # URL to access
        url = "https://bestchange.ru"

        # Perform a GET request
        driver.get(url)

        # Extract page title (or other elements)
        print(f"Page Title: {driver.title}")

        # Example: Find an element by tag
        # element = driver.find_element(By.TAG_NAME, "h1")
        print('refresh')
        time.sleep(random.uniform(10, 11))
        driver.refresh()
        time.sleep(11)
        # print(f"Heading: {element.text}")

        # Close the driver
        driver.close()
        driver.quit()
        await href_from_table()

async def get_currency(url, i):
    # await asyncio.sleep(random.uniform(5, 15))
    try:    
        timeout = aiohttp.ClientTimeout(total=60, connect=20, sock_read=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url) as response:
                page = await response.read()
        soup = BeautifulSoup(page, 'html.parser')
        try:
            div = soup.find(id="small_text")
            title = div.find("h1").text
            table = soup.find(id="content_table")
            body = table.find('tbody')
            rows = body.find_all("tr")
            pretty = soup.prettify()
            async with aiofiles.open(f"currencies2/curr_{i}.html", "w", encoding="utf-8") as f:
                await f.write(pretty)
        except Exception as ex:
            print('Handling exception associated with html: ', ex)
            # Set up Chrome options (optional)
            chrome_options = Options()
            # chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)
            # chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")

            # Automatically resolve the ChromeDriver using Selenium Manager
            driver = webdriver.Chrome(options=chrome_options)

            # URL to access
            url = "https://bestchange.ru"

            # Perform a GET request
            driver.get(url)

            # Extract page title (or other elements)
            print(f"Page Title: {driver.title}")

            # Example: Find an element by tag
            # element = driver.find_element(By.TAG_NAME, "h1")
            print('refresh')
            time.sleep(random.uniform(10, 11))
            driver.refresh()
            time.sleep(11)
            # print(f"Heading: {element.text}")

            # Close the driver
            driver.close()
            driver.quit()

            # Handliing error hrefs
            # async with aiofiles.open('error_hrefs.txt', 'a', encoding="utf-8") as f:
            #     await f.write(url)

            # Recursion
            await get_currency(url, i)

                # div = soup.find(id="small_text")
                # title = div.find("h1").text
                # table = soup.find(id="content_table")
                # body = table.find('tbody')
                # rows = body.find_all("tr")
                # for row in rows:
                #     print("Name: ", row.find(class_='ca').text)
                #     print("Rub: ", row.find(class_='bi').text)
                #     print("You will get: ", row.find_all(class_='bi')[1].text)
                #     print("Reserve: ", row.find(class_='ar arp').text)
                #     print("Reviews: ", row.find(class_='rwan').text)
                #     print("Title: ", title, '\n')
                # await writer(rows, title)
    except ConnectionTimeoutError:
        print("ConnectionTimeoutError occured")
        await get_currency(url, i)
    except Exception as ex:
        print("Exception occured: ", ex)
        await get_currency(url, i)


async def writer(rows: ResultSet, title: str, counter):
    print(counter, title)
    if not os.path.exists("main2.xlsx"):
        excel = Workbook()
        excel.save("main2.xlsx")
    excel = load_workbook("main2.xlsx")
    if title not in excel.sheetnames:
        excel.create_sheet(title)
    sheet = excel[title]
    sheet['A1'] = 'Обменник'
    sheet['B1'] = 'Отдаете'
    sheet['C1'] = 'Получаете'
    sheet['D1'] = 'Резерв'
    sheet['E1'] = 'Отзывы'
    for row, web_row in enumerate(rows, start=2):
        sheet.cell(row=row, column=1).value = web_row.find(class_='ca').text
        sheet.cell(row=row, column=2).value = web_row.find(class_='bi').text
        sheet.cell(row=row, column=3).value = web_row.find_all(class_='bi')[1].text
        sheet.cell(row=row, column=4).value = web_row.find(class_='ar arp').text
        sheet.cell(row=row, column=5).value = web_row.find(class_='rwan').text
    excel.save('main2.xlsx')


async def write_wrapper():
    ls = []

    counter = 1
    for file in os.listdir("currencies2"):
        async with aiofiles.open('currencies2/'+file, "r", encoding="utf-8") as f:
            page = await f.read()
            soup = BeautifulSoup(page, 'html.parser')
            print(file)
            try:    
                div = soup.find(id="small_text")
                title = div.find("h1").text
                table = soup.find(id="content_table")
                body = table.find('tbody')
                rows = body.find_all("tr")
                ls.append(writer(rows, title, counter))
                counter += 1
            except:
                async with aiofiles.open("invalid_hrefs.txt", 'a', encoding="utf-8") as f:
                    await f.write(file+'\n')

    await asyncio.gather(*ls)


async def main():
    main_page = 'https://www.bestchange.ru'
    await get_main_page(main_page)
    await href_from_table()
    async with aiofiles.open("hrefs2.txt", "r") as f:
        data = await f.read()
        print(type(data), data)
        all_hrefs = data.split("\n")
        sem = asyncio.Semaphore(5)

        # async def safe_get_currency(href, i):
        #     async with sem:
        #         await get_currency(href, i)
        # await asyncio.gather(*[safe_get_currency(main_page+href, i) for i, href in enumerate(all_hrefs, start=1)])



        res = len(all_hrefs) // 100
        remainder = len(all_hrefs) % 100
        chunk_size = 100

        # Calculate how many full chunks and the remainder
        res = len(all_hrefs) // chunk_size
        remainder = len(all_hrefs) % chunk_size

        # Create the chunks
        chunks = [all_hrefs[i * chunk_size:(i + 1) * chunk_size] for i in range(res)]

        # If there's a remainder, add the last chunk
        if remainder:
            chunks.append(all_hrefs[res * chunk_size:])
        print(len(chunks), )
        
        for i, ls in enumerate(chunks, start=1):
            async def safe_get_currency(href, i):
                async with sem:
                    await get_currency(href, i)
            await asyncio.gather(*[safe_get_currency(main_page+href, i) for i, href in enumerate(ls, start=1)])
            await write_wrapper()
            print(f"{i} hundred saved to the file")
            print(f"Removing: ")
            for f in os.listdir('currencies2'):
                os.remove('currencies2/'+f)
            print(f"removed.")
            print("Sleep time for 60 seconds")
            await asyncio.sleep(60)


    # await write_wrapper()


if __name__ == "__main__":
    asyncio.run(main())

    #     # excel = Workbook(write_only=True)
    # excel = load_workbook('main2.xlsx')
    # # if 'Osnovnoy2' not in excel.sheetnames:
    # #     excel.create_sheet("Osnovnoy2")
    # for i, sheet in enumerate(excel.sheetnames):
    #     print(f"Sheet {i}: ", sheet)
    # #     del excel[sheet]
    #     # break
    # # del excel['Osnovnoy']
    # excel.active = excel.sheetnames.index("Osnovnoy2")
    # a = excel.active
    # a['A1'] = 2
    # # print(a.columns)
    # # for row in a.rows:
    # #     for cell in row:
    # #         cell.value = ''
    # for row in a.iter_cols(min_row=1, min_col=1):
    #     for i, cell in enumerate(row):
    #         cell.value = f"Updated{i}"
    # excel.save('main2.xlsx')



# async def x(i):
#     # print(f"start {i}")
#     # await asyncio.sleep(1)
#     # print(f"end {i}")
#     return i

# # run x(0)..x(10) concurrently and process results as they arrive
# async def y():
#     # start = time.time()
#     # for f in asyncio.as_completed([x(i) for i in range(10000)]):
#         # result = await f
#     # end = time.time()
#     start = time.time()
#     res = await asyncio.gather(*[x(i) for i in range(10000)])
#     end = time.time()
#     print(end-start)

# async def main():
#     await y()

# asyncio.run(main())





# import time
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options

# # Set up Chrome options (optional)
# chrome_options = Options()
# # chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)
# # chrome_options.add_argument("--disable-gpu")
# chrome_options.add_argument("--no-sandbox")

# # Automatically resolve the ChromeDriver using Selenium Manager
# driver = webdriver.Chrome(options=chrome_options)

# # URL to access
# url = "https://bestchange.ru"

# # Perform a GET request
# driver.get(url)

# # Extract page title (or other elements)
# print(f"Page Title: {driver.title}")

# # Example: Find an element by tag
# element = driver.find_element(By.TAG_NAME, "h1")
# print('refresh')
# driver.refresh()
# time.sleep(20)
# print(f"Heading: {element.text}")

# # Close the driver
# driver.quit()
