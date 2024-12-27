import json
import os
import asyncio
from itertools import permutations
import re
import time

import aiofiles
import aiohttp
from bs4 import BeautifulSoup
from numpy import save
from openpyxl import Workbook, load_workbook

API = '57e38271ed710678cffd0e0ddcdcb278'

url = 'https://www.bestchange.app/v2'

dir = 'main2'
dir_2 = 'change_rates'

MAX_REQUESTS = 300

semaphore = asyncio.Semaphore(MAX_REQUESTS)

# perm = permutations(range(10), 2)
# com = combinations(range(10), 2)
# pro = product(range(5), repeat=1)
# for p in pro:
#     print(p)



async def get_main_page(url):
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            read = await response.read()
            soup = BeautifulSoup(read, "html.parser")
            # print(read)
            try:    
                table = soup.find(id="curr_tab_c")
                body = table.find("tbody")
                pretty = soup.prettify()
            except Exception as e:
                print("Error handled on main page: ", e)
                await get_main_page(url)

    async with aiofiles.open('index.html', "w", encoding="utf-8") as f:
        await f.write(pretty)


async def href_from_table() -> None:
    async with aiofiles.open("index.html", "r", encoding="utf-8") as f:
        html = await f.read()
    soup = BeautifulSoup(html, 'html.parser')
    top = soup.find(id="curr_top")
    ass1 = top.find_all("a")
    top_hrefs = [href.get('href') for href in ass1]
    table = soup.find(id="curr_tab_c")
    body = table.find("tbody")
    ass2 = body.find_all("a")
    hrefs = [href.get("href") for href in ass2]
    # print(len(hrefs))
    # with open("_a_tag.html", "w", encoding='utf-8') as f:
    #     f.write('\n'.join([str(tag) for tag in ass]))
    # for raw in ass:
    #     print(raw)
    async with aiofiles.open("hrefs.txt", "w", encoding='utf-8') as f:
        await f.write('\n'.join(top_hrefs))
        await f.write('\n'+'\n'.join(hrefs))
    

async def filter_currencies(currencies: dict) -> dict:
    # print(currencies)
    cur_data: list[dict] = currencies.get("currencies")
    async with aiofiles.open("hrefs.txt", 'r', encoding="utf8") as f:
        data = await f.readlines()
    dc = {}
    dc['currencies'] = []
    ls_id_from = []
    ls_id_to = []
    for line in data:
        # print(line)
        st_line = re.sub(r"[/.]|html|\n", '', line)
        # print(st_line)
        ls = re.split(r"-to-", st_line)
        # print(ls)
        from_, to_ = ls
        # print(from_)
        # filtered_cur = [cur for cur in cur_data if (cur['urlname'] == from_ or cur['urlname'] == to_) and cur['id'] not in [c.get('id') for c in dc['currencies']]]
        id_from = [[cur['id'], cur['urlname']] for cur in cur_data if cur.get('urlname') == from_]
        id_to = [[cur['id'], cur['urlname']] for cur in cur_data if cur.get('urlname') == to_]
        ls_id_from.extend(id_from)
        ls_id_to.extend(id_to)
        # dc['currencies'].extend(filtered_cur)
    await asyncio.gather(*[get_rates(url, id_from[0], id_to[0], id_from[1], id_to[1]) for id_from, id_to in zip(ls_id_from, ls_id_to)])
    return dc


async def get_currencies(dir: str, url: str) -> None:
    async with aiohttp.ClientSession() as session:
        async with session.get(f"{url}/{API}/currencies/ru") as response:
            currencies = await response.json()
            await filter_currencies(currencies)
    if not os.path.exists('main2'):
        os.mkdir('main2')
    async with aiofiles.open(f"{dir}/currencies.json", "w", encoding="utf-8") as f:
        await f.write(json.dumps(currencies, indent=4, ensure_ascii=False))
    
    return currencies


async def get_changers(dir: str, url: str) -> None:
    async with aiohttp.ClientSession() as sesssion:
        async with sesssion.get(f'{url}/{API}/changers/ru') as response:
            changers = await response.json()
    async with aiofiles.open(f"{dir}/changers.json", "w", encoding='utf-8') as f:
        await f.write(json.dumps(changers, indent=4, ensure_ascii=False))

    return changers


# async def get_combinationOfcurrencies(dir: str, url: str) -> None:
#     currencies = await get_currencies(dir, url)
#     changers = await get_changers(dir, url)
#     # print(currencies['currencies'])
#     id_currencies = [cur['id'] for cur in currencies["currencies"]]
#     id_chamgers = [chan['id'] for chan in changers["changers"]]

#     combinations = permutations(id_currencies, 2)
#     # print(id_currencies, len(id_currencies))
#     ls = [com for com in combinations]
#     # print('\n', ls, len(ls))

#     return ls


async def get_rates(url: str, from_, to_, from_name, to_name):
    try:
        total = aiohttp.ClientTimeout(total=30, connect=20, sock_read=30)
        if not os.path.exists('change_rates'):
            os.mkdir("change_rates")
        async with semaphore:
            async with aiohttp.ClientSession(timeout=total) as session:
                async with session.get(f"{url}/{API}/rates/{from_}-{to_}") as response:
                    if response.status == 429:
                        print(f"Rate limited: {from_}-{to_}, retrying after delay...")
                        await asyncio.sleep(2)  # Wait before retrying
                        await get_rates(url, from_, to_, from_name, to_name)
                    rates = await response.json()
        if rates['rates'][f'{from_}-{to_}']:
            print(f'{from_}-{to_}')
            async with aiofiles.open(f"change_rates/{from_name}-to-{to_name}.json", "w", encoding='utf-8') as f:
                await f.write(json.dumps(rates, indent=4, ensure_ascii=False))
    except Exception as e:
        print('Error handled', e)
        await get_rates(url, from_, to_, from_name, to_name)


def write_to_excel(title: str, ):
    pass
    # TODO need to finish    


async def save_to_excel(dir: str):
    for file in os.listdir(dir):
        async with aiofiles.open(f'{dir}/{file}', 'r', encoding='utf-8') as file:
            data = await file.read()
            data = json.loads(data)
            rates = data.get("rates")
            for key, value in rates.items():
                from_, to_ = key.split("-")
                from_, to_ =int(from_),int(to_)
                async with aiofiles.open(f"main2/currencies.json", 'r', encoding="utf-8") as f:
                    currencies = await f.read()
                    currencies = json.loads(currencies)
                    data = currencies.get("currencies")
                    fr = [d['name'] for d in data if d['id'] == from_]
                    to = [d['name'] for d in data if d['id'] == to_]
                    title = f"{fr[0]} на {to[0]}"
                    print(title)
                if not os.path.exists(f"main.xlsx"):
                    excel = Workbook()
                    excel.save(f"main.xlsx")
                
                
                    # print(f"From: {fr[0]}\nTo: {to[0]}")
        # break
    # if not os.path.exists("main_3.xlsx"):
    #     excel = Workbook()
    #     excel.save("main_3.xlsx")
    # excel = load_workbook("main_3.xlsx")
    # if title not in excel.sheetnames:
    #     excel.create_sheet(title)
    # sheet = excel[title]
    # sheet['A1'] = 'Обменник'
    # sheet['B1'] = 'Отдаете'
    # sheet['C1'] = 'Получаете'
    # sheet['D1'] = 'Резерв'
    # sheet['E1'] = 'Отзывы'
    # for row, web_row in enumerate(rows, start=2):
    #     sheet.cell(row=row, column=1).value = web_row.find(class_='ca').text
    #     sheet.cell(row=row, column=2).value = web_row.find(class_='bi').text
    #     sheet.cell(row=row, column=3).value = web_row.find_all(class_='bi')[1].text
    #     sheet.cell(row=row, column=4).value = web_row.find(class_='ar arp').text
    #     sheet.cell(row=row, column=5).value = web_row.find(class_='rwan').text
    # excel.save('main2.xlsx')


async def main():
    main_page = 'https://www.bestchange.ru'
    # await get_main_page(main_page)
    # await href_from_table()
    # await get_currencies(dir, url)
    # combinations = await get_combinationOfcurrencies(dir, url)
    # await asyncio.gather(*[get_rates(url, from_, to_) for from_, to_ in combinations])
    # # await get_rates(url, 48, 213)
    # print(len([f for f in os.listdir('change_rates')]))
    await save_to_excel(dir_2)


if __name__ == "__main__":
    asyncio.run(main())
